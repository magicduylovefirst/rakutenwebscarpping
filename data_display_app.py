import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import json
import os
import csv
import numpy as np # For random data in placeholder

# Attempt to import actual processors, fall back to placeholders if not found
try:
    from excel_processor import load_excel_data as actual_load_excel_data, \
                                get_orange_rows_info as actual_get_orange_rows_info, \
                                export_to_csv as actual_export_to_csv
except ImportError:
    actual_load_excel_data = None
    actual_get_orange_rows_info = None
    actual_export_to_csv = None

try:
    from rakuten_item_fetcher import fetch_item_data as actual_fetch_item_data
except ImportError:
    actual_fetch_item_data = None

# --- Placeholder functions if actual modules are not available ---
def placeholder_load_excel_data(filepath):
    try:
        df = pd.read_excel(filepath, engine='openpyxl')
        # If 'SKU' column doesn't exist, try to create one from index for identification
        if 'SKU' not in df.columns and not df.empty:
            df['SKU_generated_id'] = df.index.astype(str)
        return df
    except FileNotFoundError:
        messagebox.showerror("Error", f"Excel file not found: {filepath}")
        return pd.DataFrame()
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load Excel file '{filepath}': {e}")
        return pd.DataFrame()

def placeholder_get_orange_rows_info(filepath):
    orange_indices = []
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(filepath)
        sheet = workbook.active # Assuming the first sheet

        # Common orange ARGB hex codes. Excel's "orange" can vary.
        TARGET_ORANGE_RGB = "FFFF9900" 

        for r_idx, row in enumerate(sheet.iter_rows(min_row=1)): # r_idx will be 0-based
            is_orange_row = False
            for cell in row:
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    if cell.fill.start_color.rgb == TARGET_ORANGE_RGB:
                        is_orange_row = True
                        break 
            if is_orange_row:
                orange_indices.append(r_idx)
    except Exception as e:
        print(f"Error reading orange rows from '{filepath}': {e}. Assuming no orange rows.")
    return orange_indices

def placeholder_export_to_csv(excel_filepath, output_csv_filepath):
    try:
        df = pd.read_excel(excel_filepath, engine='openpyxl')
        df.to_csv(output_csv_filepath, index=False, quoting=csv.QUOTE_ALL)
        return True
    except Exception as e:
        messagebox.showerror("Error", f"Failed to export to CSV: {e}")
        return False

def placeholder_fetch_item_data(sku_list):
    print(f"Placeholder: Fetching data for SKUs: {sku_list}")
    updated_data = []
    for sku in sku_list:
        updated_data.append({
            'SKU': sku, 
            '商品名': f'Updated Name for {sku} ({np.random.choice(["A", "B"])})', 
            '価格': float(np.random.randint(1000, 9000)),
            '在庫': int(np.random.randint(0, 50)),
            '備考': f'Recalculated {pd.Timestamp.now().strftime("%H:%M")}'
        })
    return pd.DataFrame(updated_data)

# Use actual functions if available, otherwise use placeholders
load_excel_data = actual_load_excel_data if actual_load_excel_data else placeholder_load_excel_data
get_orange_rows_info = actual_get_orange_rows_info if actual_get_orange_rows_info else placeholder_get_orange_rows_info
export_to_csv_file = actual_export_to_csv if actual_export_to_csv else placeholder_export_to_csv
fetch_item_data = actual_fetch_item_data if actual_fetch_item_data else placeholder_fetch_item_data

ARAKI_XLSX_PATH = 'araki.xlsx' 
PREVIOUS_DATA_PATH = 'previous_data.json'

class ExcelApp:
    def __init__(self, root_window):
        self.root = root_window
        self.root.title("Excel Data Viewer")
        self.root.geometry("1000x700")

        self.data_df = pd.DataFrame()
        self.previous_data_df = pd.DataFrame()
        self.orange_row_indices = [] # Store 0-based indices of orange rows
        self.sku_column_name = 'SKU' # Default SKU column name

        self._setup_menu()
        self._setup_toolbar()
        self._setup_sku_input()
        self._setup_table()

        self.load_persistent_data()
        self.load_initial_data() # This will also determine self.sku_column_name

    def _determine_sku_column(self, df):
        if 'SKU' in df.columns:
            return 'SKU'
        elif 'SKU_generated_id' in df.columns:
            return 'SKU_generated_id'
        elif not df.empty:
            return df.columns[0]
        return None

    def _setup_menu(self):
        menubar = tk.Menu(self.root)
        filemenu = tk.Menu(menubar, tearoff=0)
        filemenu.add_command(label="Export araki.xlsx to CSV", command=self.export_araki_to_csv)
        filemenu.add_separator()
        filemenu.add_command(label="Exit", command=self.on_closing)
        menubar.add_cascade(label="File", menu=filemenu)
        helpmenu = tk.Menu(menubar, tearoff=0)
        helpmenu.add_command(label="About", command=self.show_about)
        menubar.add_cascade(label="Help", menu=helpmenu)
        self.root.config(menu=menubar)

    def _setup_toolbar(self):
        toolbar = tk.Frame(self.root, bd=1, relief=tk.RAISED)
        recalc_button = ttk.Button(toolbar, text="Recalculate All", command=self.recalculate_all)
        recalc_button.pack(side=tk.LEFT, padx=2, pady=2)
        recalc_selected_button = ttk.Button(toolbar, text="Selected Recalculate", command=self.recalculate_selected)
        recalc_selected_button.pack(side=tk.LEFT, padx=2, pady=2)
        export_button = ttk.Button(toolbar, text="CSV Export (araki.xlsx)", command=self.export_araki_to_csv)
        export_button.pack(side=tk.LEFT, padx=2, pady=2)
        toolbar.pack(side=tk.TOP, fill=tk.X)

    def _setup_sku_input(self):
        sku_frame = tk.Frame(self.root)
        tk.Label(sku_frame, text="Input SKU for Selected Recalculate:").pack(side=tk.LEFT, padx=5, pady=5)
        self.sku_entry = ttk.Entry(sku_frame, width=30)
        self.sku_entry.pack(side=tk.LEFT, padx=5, pady=5)
        sku_frame.pack(side=tk.TOP, fill=tk.X)

    def _setup_table(self):
        frame = tk.Frame(self.root)
        frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.tree = ttk.Treeview(frame, show='headings')
        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side='right', fill='y')
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        hsb.pack(side='bottom', fill='x')
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.pack(fill=tk.BOTH, expand=True)

        self.tree.tag_configure('changed', background='yellow')
        self.tree.tag_configure('new_item', background='lightgreen')
        self.tree.tag_configure('orange_excel', background='orange')

    def load_initial_data(self):
        self.data_df = load_excel_data(ARAKI_XLSX_PATH)
        self.sku_column_name = self._determine_sku_column(self.data_df)

        if not self.data_df.empty:
            self.orange_row_indices = get_orange_rows_info(ARAKI_XLSX_PATH)
        else:
            if os.path.exists(ARAKI_XLSX_PATH):
                messagebox.showwarning("Data Load", f"Could not load data from {ARAKI_XLSX_PATH}. It might be empty or corrupted.")

        self.update_table_display(highlight_new=True)
        if self.previous_data_df.empty and not self.data_df.empty:
            self.previous_data_df = self.data_df.copy()
            self.save_persistent_data()

    def load_persistent_data(self):
        if os.path.exists(PREVIOUS_DATA_PATH):
            try:
                self.previous_data_df = pd.read_json(PREVIOUS_DATA_PATH, orient='split')
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load persistent data from {PREVIOUS_DATA_PATH}: {e}")
                self.previous_data_df = pd.DataFrame()
        else:
            self.previous_data_df = pd.DataFrame()

    def save_persistent_data(self):
        try:
            if isinstance(self.data_df, pd.DataFrame) and not self.data_df.empty:
                self.data_df.to_json(PREVIOUS_DATA_PATH, orient='split', indent=4)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save persistent data to {PREVIOUS_DATA_PATH}: {e}")

    def update_table_display(self, highlight_new=False, skus_just_changed=None):
        for i in self.tree.get_children():
            self.tree.delete(i)
        
        if self.data_df.empty:
            self.tree["columns"] = []
            self.tree["displaycolumns"] = []
            return

        self.tree["columns"] = list(self.data_df.columns)
        self.tree["displaycolumns"] = list(self.data_df.columns)
        
        for col in self.data_df.columns:
            self.tree.heading(col, text=col, command=lambda _col=col: self.sort_column(_col, False))
            self.tree.column(col, anchor=tk.W, width=120)

        for df_idx, row_data in self.data_df.iterrows():
            values = list(row_data)
            item_iid = str(df_idx)
            
            tags_to_apply = []

            if df_idx in self.orange_row_indices:
                tags_to_apply.append('orange_excel')

            current_sku = row_data[self.sku_column_name] if self.sku_column_name and self.sku_column_name in row_data else None

            if highlight_new and current_sku and not self.previous_data_df.empty and self.sku_column_name in self.previous_data_df.columns:
                if current_sku not in self.previous_data_df[self.sku_column_name].values:
                    tags_to_apply.append('new_item')
            
            if skus_just_changed and current_sku in skus_just_changed:
                if not self.previous_data_df.empty and self.sku_column_name in self.previous_data_df.columns:
                    prev_row_series = self.previous_data_df[self.previous_data_df[self.sku_column_name] == current_sku]
                    if not prev_row_series.empty:
                        prev_row = prev_row_series.iloc[0]
                        if not row_data.equals(prev_row.reindex(row_data.index).fillna(pd.NA)):
                            tags_to_apply.append('changed')
            
            self.tree.insert("", tk.END, values=values, iid=item_iid, tags=tuple(tags_to_apply))

    def sort_column(self, col, reverse):
        try:
            data_list = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
            try:
                data_list.sort(key=lambda t: float(t[0]), reverse=reverse)
            except ValueError:
                data_list.sort(key=lambda t: str(t[0]).lower(), reverse=reverse)

            for index, (val, k) in enumerate(data_list):
                self.tree.move(k, '', index)
            
            self.tree.heading(col, command=lambda _col=col: self.sort_column(_col, not reverse))
        except Exception as e:
            print(f"Error sorting column {col}: {e}")

    def _internal_recalculate(self, skus_to_process):
        if not skus_to_process:
            messagebox.showinfo("Info", "No SKUs provided for recalculation.")
            return

        if not self.sku_column_name or self.sku_column_name not in self.data_df.columns:
            messagebox.showerror("Error", "SKU column not configured or found in data.")
            return

        self.previous_data_df = self.data_df.copy()
        fetched_df = fetch_item_data(skus_to_process)

        if fetched_df.empty or self.sku_column_name not in fetched_df.columns:
            messagebox.showwarning("Recalculate", "No data returned from fetch operation or SKU column missing in fetched data.")
            return

        original_index_name = self.data_df.index.name
        self.data_df = self.data_df.set_index(self.sku_column_name, drop=False)
        
        for _, fetched_row in fetched_df.iterrows():
            sku = fetched_row[self.sku_column_name]
            if sku in self.data_df.index:
                for col in fetched_df.columns:
                    if col in self.data_df.columns:
                        self.data_df.loc[sku, col] = fetched_row[col]

        self.data_df = self.data_df.reset_index(drop=True)
        if original_index_name:
             self.data_df.index.name = original_index_name

        self.update_table_display(skus_just_changed=skus_to_process)
        self.save_persistent_data()
        messagebox.showinfo("Recalculate", f"Recalculation complete for: {', '.join(skus_to_process)}.")

    def recalculate_all(self):
        if self.data_df.empty or not self.sku_column_name:
            messagebox.showinfo("Info", "No data or SKU column to recalculate.")
            return
        all_skus = self.data_df[self.sku_column_name].unique().tolist()
        self._internal_recalculate(all_skus)

    def recalculate_selected(self):
        selected_iids = self.tree.selection()
        skus_to_recalculate = []

        if not self.sku_column_name:
            messagebox.showerror("Error", "SKU column not identified.")
            return

        if selected_iids:
            try:
                selected_df_indices = [int(iid) for iid in selected_iids]
                valid_indices = [idx for idx in selected_df_indices if idx in self.data_df.index]
                if valid_indices:
                    skus_to_recalculate = self.data_df.loc[valid_indices, self.sku_column_name].unique().tolist()
            except ValueError:
                messagebox.showerror("Error", "Selection contains non-integer DataFrame indices.")
                return
        
        if not skus_to_recalculate:
            input_sku = self.sku_entry.get().strip()
            if input_sku:
                if self.sku_column_name in self.data_df.columns and \
                   input_sku in self.data_df[self.sku_column_name].values:
                    skus_to_recalculate = [input_sku]
                else:
                    messagebox.showinfo("Info", f"SKU '{input_sku}' not found in table.")
                    return
            else:
                messagebox.showinfo("Info", "No items selected in table and no SKU entered in input field.")
                return
        
        if not skus_to_recalculate:
            messagebox.showinfo("Info", "No SKUs identified for recalculation.")
            return

        self._internal_recalculate(skus_to_recalculate)

    def export_araki_to_csv(self):
        if not os.path.exists(ARAKI_XLSX_PATH):
            messagebox.showerror("Error", f"{ARAKI_XLSX_PATH} not found. Cannot export.")
            return
            
        save_path = filedialog.asksaveasfilename(defaultextension=".csv",
                                                 filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                                                 title="Save araki.xlsx content as CSV")
        if not save_path:
            return
        
        if export_to_csv_file(ARAKI_XLSX_PATH, save_path):
            messagebox.showinfo("Export Successful", f"Data from {ARAKI_XLSX_PATH} exported to {save_path}")

    def show_about(self):
        messagebox.showinfo("About", "Excel Data Viewer\nVersion 1.0\n\nHandles display and processing of Excel data related to Rakuten items.")

    def on_closing(self):
        self.save_persistent_data()
        self.root.destroy()

if __name__ == '__main__':
    main_root = tk.Tk()
    app = ExcelApp(main_root)
    main_root.protocol("WM_DELETE_WINDOW", app.on_closing)
    main_root.mainloop() 
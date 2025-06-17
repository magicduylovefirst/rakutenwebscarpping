import requests
import json
import time
import re
import pandas as pd
from typing import List, Dict
from datetime import datetime
import base64
import openpyxl

# Constants
API_ENDPOINT = 'https://app.rakuten.co.jp/services/api/IchibaItem/Search/20220601'
APP_ID = "1085274442429696242"  # Rakuten Application ID

# Define shop configurations
SHOPS = {
    'waste': {  # e-life＆work shop
        'name': 'e-life＆work shop',
        'shop_code': 'waste',
        'base_url': 'https://item.rakuten.co.jp/waste/',
        'description': 'Main shop for safety equipment and work gear',
        'sku_format': lambda sku: sku.split('-')[1]  # Extract model number (e.g., 1271a029)
    },
    'kougushop': {  # 工具ショップ
        'name': '工具ショップ',
        'shop_code': 'kougushop',
        'base_url': 'https://item.rakuten.co.jp/kougushop/',
        'description': 'Specialized in tools and safety equipment',
        'sku_format': lambda sku: f"{sku.split('-')[1]}-{sku.split('-')[2]}"  # Format: 1271a029-025
    },
    'kouei-sangyou': {  # 晃栄産業
        'name': '晃栄産業',
        'shop_code': 'kouei-sangyou',
        'base_url': 'https://item.rakuten.co.jp/kouei-sangyou/',
        'description': 'Industrial safety equipment supplier',
        'sku_format': lambda sku: f"fcp209"  # Fixed format for CP209
    },
    'dear-worker': {  # dear-worker
        'name': 'dear-worker',
        'shop_code': 'dear-worker',
        'base_url': 'https://item.rakuten.co.jp/dear-worker/',
        'description': 'Worker safety equipment specialist',
        'sku_format': lambda sku: f"cp209boa"  # Fixed format for CP209 BOA
    }
}

def truncate_str(s: str, length: int) -> str:
    """Truncate string to specified length, adding ... if truncated"""
    if len(s) <= length:
        return s
    return s[:length-3] + "..."

def clean_text(text):
    """Clean text by removing extra whitespace and normalizing newlines"""
    if not text:
        return ""
    # Replace multiple spaces with single space
    text = re.sub(r'\s+', ' ', text)
    # Replace multiple newlines with single newline
    text = re.sub(r'\n+', '\n', text)
    return text.strip()

def extract_sizes(caption):
    """Extract size information from item caption"""
    sizes = []
    if not caption:
        return sizes
        
    # Look for size patterns like "25.5cm" or "25.5 cm"
    size_pattern = r'(\d{2}\.?\d?)\s*cm'
    matches = re.finditer(size_pattern, caption)
    for match in matches:
        size = match.group(1)
        if 20 <= float(size) <= 31:  # Reasonable shoe size range
            sizes.append(f"{size}cm")
    return sorted(list(set(sizes)))  # Remove duplicates and sort

def extract_specs(caption):
    """Extract specifications from item caption"""
    specs = {}
    if not caption:
        return specs
    
    # Common Japanese spec patterns
    known_specs = {
        '幅/ラスト': r'幅[/／]?ラスト[：:]?\s*([^。\n]+)',
        'アッパー素材': r'アッパー素材[：:]?\s*([^。\n]+)',
        'アウター素材': r'アウター素材[：:]?\s*([^。\n]+)',
        'インナーソール': r'インナーソール[：:]?\s*([^。\n]+)',
        '品番': r'品番\s*[：:]\s*([^。\n]+)',
        'サイズ': r'サイズ[：:]?\s*([^。\n]+)',
        '重量': r'重量[：:]?\s*([^。\n]+)',
        '生産国': r'Made in ([^。\n]+)',
    }
    
    # Extract each specification
    for key, pattern in known_specs.items():
        match = re.search(pattern, caption)
        if match:
            value = match.group(1).strip()
            # Clean up the value
            value = re.sub(r'\s+', ' ', value)  # Normalize spaces
            value = re.sub(r'[：:]\s*', ': ', value)  # Normalize colons
            specs[key] = value
    
    return specs

def format_sku_for_shop(sku: str, shop_info: dict) -> str:
    """Format SKU based on shop's pattern"""
    try:
        return shop_info['sku_format'](sku)
    except Exception as e:
        print(f"Error formatting SKU {sku} for {shop_info['name']}: {e}")
        return sku

def fetch_rms_details(manage_number: str) -> dict:
    """Fetch detailed product info from RMS API using manageNumber"""
    try:
        # Your credentials
        service_secret = "SP208989_Rj1XXyKcXKwL813w"
        license_key = "SL208989_j675LcBF0x444bE6"
        
        # Create Authorization header
        raw_credential = f"{service_secret}:{license_key}"
        encoded_credential = base64.b64encode(raw_credential.encode()).decode()
        authorization_header = f"ESA {encoded_credential}"
        
        # API endpoint
        url = f"https://api.rms.rakuten.co.jp/es/2.0/items/manage-numbers/{manage_number}"
        headers = {
            'Authorization': authorization_header,
            'Content-Type': 'application/json'
        }

        print(f"Fetching RMS data for manageNumber: {manage_number}")
        response = requests.get(url, headers=headers)
        
        if response.status_code == 404:
            print(f"Item not found in RMS API for manageNumber: {manage_number}")
            return {
                'title': 'N/A',
                'reference_price': '-',
                'standard_price': 0
            }
            
        response.raise_for_status()
        rms_data = response.json()
        
        # Get first variant (if exists) for pricing
        variant_data = list(rms_data.get("variants", {}).values())[0] if rms_data.get("variants") else {}
        reference_price = variant_data.get("referencePrice", {}).get("value") or '-'
        standard_price = variant_data.get("standardPrice", 0)
        
        return {
            'title': rms_data.get('title', 'N/A'),
            'reference_price': reference_price,
            'standard_price': standard_price
        }
    except Exception as e:
        print(f"[RMS API ERROR] manageNumber={manage_number}: {e}")
        return {
            'title': 'N/A',
            'reference_price': '-',
            'standard_price': 0
        }

def fetch_ichiba_details(code: str, shop_code: str, shop_info: dict) -> list:
    """Fetch shop-specific info from Ichiba API"""
    try:
        search_code = format_sku_for_shop(code, shop_info)
        params = {
            'applicationId': APP_ID,
            'shopCode': shop_code,
            'keyword': search_code,
            'hits': 10,
            'format': 'json',
            'availability': 1
        }
        
        print(f"Searching in Ichiba API for {shop_info['name']} with code: {search_code}")
        response = requests.get(API_ENDPOINT, params=params)
        response.raise_for_status()
        data = response.json()
        
        shop_items = []
        if data.get('Items'):
            for item_data in data['Items']:
                item = item_data['Item']
                price = item.get('itemPrice', 0)
                
                shop_items.append({
                    'manage_number': item.get('itemCode', 'N/A').replace(f"{shop_code}:", ''),
                    'price': price,
                    'points': item.get('points', 0),
                    'coupon': item.get('couponPrice', 0),
                    'availability': item.get('availability', 0),
                    'url': item.get('itemUrl', ''),
                    'name': item.get('itemName', 'N/A')
                })
        return shop_items
    except Exception as e:
        print(f"[Ichiba API ERROR] code={code}, shop={shop_code}: {e}")
        return []

def fetch_item_details(code: str) -> dict:
    """Fetch detailed item information from multiple shops"""
    items_by_sku = {}  # Dictionary to group items by SKU
    tax_rate = 1.1
    
    # Fetch data from each shop using Ichiba API
    for shop_code, shop_info in SHOPS.items():
        ichiba_items = fetch_ichiba_details(code, shop_code, shop_info)
        
        for ichiba_item in ichiba_items:
            manage_number = ichiba_item['manage_number']
            price = ichiba_item['price']
            
            # Get RMS data using manageNumber from Ichiba
            rms_data = fetch_rms_details(manage_number)
            standard_price = rms_data['standard_price']
            fa_price_ex_tax = int(int(standard_price) / tax_rate) if standard_price else int(price / tax_rate)
            
            # Create or update item entry
            if code not in items_by_sku:
                items_by_sku[code] = {
                    'original_sku': code,
                    'search_code_used': format_sku_for_shop(code, shop_info),
                    'product_info': {
                        '商品管理番号': manage_number,
                        '商品名':  ichiba_item['name'],
                        '検索条件': code,
                        '検索除外': '-',
                        '在庫': '○' if ichiba_item['availability'] == 1 else '×',
                        '定価': rms_data['reference_price'],
                        '仕入金額': '-',
                        '平均単価': '-',
                        'FA売価(税抜)': fa_price_ex_tax,
                        '粗利': '-',
                        'RT後の利益': '-',
                        'FA売価(税込)': price
                    },
                    'shop_info': {}  # Will hold multiple shop entries
                }
            
            # Add shop-specific info
            items_by_sku[code]['shop_info'][shop_code] = {
                'shop_name': shop_info['name'],
                'shop_code': shop_code,
                '価格': price,
                'ポイント': ichiba_item['points'],
                'クーポン': ichiba_item['coupon'],
                '在庫状況': '在庫あり' if ichiba_item['availability'] == 1 else '在庫なし',
                'URL': ichiba_item['url']
            }
            
            # Debug output
            print(f"\nProcessed item:")
            print(f"SKU: {code}")
            print(f"Shop: {shop_info['name']}")
            print(f"ManageNumber: {manage_number}")
            print(f"RMS Title: {rms_data['title']}")
            print(f"Ichiba Price: {price}")
            print(f"RMS Standard Price: {standard_price}")
            print(f"FA Price (excl. tax): {fa_price_ex_tax}")
    
    # Convert dictionary to list
    all_items = list(items_by_sku.values())
    return {'items': all_items}


def read_skus_from_excel(excel_path):
    """Read SKUs from first column of Excel file"""
    try:
        df = pd.read_excel(excel_path)
        # Clean SKUs - remove header row if it exists
        skus = df.iloc[:, 0].dropna().astype(str).tolist()
        return [sku for sku in skus if sku.lower() != 'skuコード']
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []

def print_table_header(columns: List[str], widths: List[int]) -> None:
    """Print table header with proper formatting"""
    # Print top border
    print("+" + "+".join("-" * (w + 2) for w in widths) + "+")
    
    # Print column headers
    row = "|"
    for col, width in zip(columns, widths):
        row += f" {col.center(width)} |"
    print(row)
    
    # Print separator
    print("+" + "+".join("=" * (w + 2) for w in widths) + "+")

def print_table_row(data: List[str], widths: List[int]) -> None:
    """Print a table row with proper formatting"""
    # Split long data into multiple rows if needed
    max_lines = max(len(str(d).split('\n')) for d in data)
    
    for line_num in range(max_lines):
        row = "|"
        for value, width in zip(data, widths):
            value_str = str(value).split('\n')[line_num] if line_num < len(str(value).split('\n')) else ''
            value_str = truncate_str(value_str, width)
            
            # Right-align numbers (including those with ¥), left-align text
            if value_str.startswith("¥") or value_str.replace(",", "").isdigit():
                row += f" {value_str.rjust(width)} |"
            else:
                row += f" {value_str.ljust(width)} |"
        print(row)
    
    # Print bottom border
    print("+" + "+".join("-" * (w + 2) for w in widths) + "+")

def main():
    start_time = time.time()
    
    # Read SKUs from Excel
    excel_path = "New folder/araki.xlsx"
    skus = read_skus_from_excel(excel_path)
    
    if not skus:
        print("No SKUs found in Excel file")
        return
        
    print(f"Processing {len(skus)} SKUs...")
    print(f"Start time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("\nSearching in the following shops:")
    for shop_code, shop_info in SHOPS.items():
        print(f"- {shop_info['name']} ({shop_code})")
        print(f"  URL: {shop_info['base_url']}")
        print(f"  Description: {shop_info['description']}\n")
    
    # Process all SKUs
    all_results = {
        'items': [],
        'shops_searched': {
            shop_code: {
                'name': info['name'],
                'base_url': info['base_url'],
                'description': info['description']
            } for shop_code, info in SHOPS.items()
        }
    }
    processed_count = 0
    items_per_shop = {shop_code: 0 for shop_code in SHOPS.keys()}
    
    for sku in skus:
        start_item_time = time.time()
        print(f"\nFetching data for SKU: {sku}")
        result = fetch_item_details(sku)
        
        # Count items per shop
        for item in result['items']:
            for shop_info in item['shop_info'].values():
                shop_code = shop_info['shop_code']
                if shop_code in items_per_shop:
                    items_per_shop[shop_code] += 1
        
        all_results['items'].extend(result['items'])
        processed_count += len(result['items'])
        
        # Calculate and show progress
        elapsed_item_time = time.time() - start_item_time
        print(f"Time for this SKU: {elapsed_item_time:.2f} seconds")
        time.sleep(1)  # Be nice to Rakuten's servers
    
    # Calculate total time
    total_time = time.time() - start_time
    
    # Add timing info to results
    all_results['metadata'] = {
        'total_time_seconds': round(total_time, 2),
        'average_time_per_sku': round(total_time / len(skus), 2) if skus else 0,
        'total_skus_processed': len(skus),
        'total_items_found': processed_count,
        'items_per_shop': items_per_shop,
        'start_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    # Save results to JSON file
    with open('results.json', 'w', encoding='utf-8') as f:
        json.dump(all_results, ensure_ascii=False, indent=2, fp=f)
    
    print(f"\nProcessing completed:")
    print(f"Total time: {total_time:.2f} seconds")
    print(f"Average time per SKU: {total_time / len(skus):.2f} seconds")
    print(f"Total SKUs processed: {len(skus)}")
    print(f"Items found per shop:")
    for shop_code, count in items_per_shop.items():
        print(f"- {SHOPS[shop_code]['name']}: {count} items")
    print(f"Results saved to results.json")

def update_excel_with_results(json_path, excel_path):
    """Update Excel file with results from JSON data"""
    # Read JSON data
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Load Excel file
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Start from row 4 (as per previous pattern)
    current_row = 4
    
    # Process each item
    for item in data['items']:
        product_info = item['product_info']
        shop_info = item['shop_info']
        
        # Update common fields
        if product_info.get('商品名'):
            ws[f'C{current_row}'] = product_info['商品名']
            
        if product_info.get('商品管理番号'):
            ws[f'B{current_row}'] = product_info['商品管理番号']
            
        if product_info.get('検索条件'):
            ws[f'D{current_row}'] = product_info['検索条件']
            
        # Update price and tax fields
        if product_info.get('FA売価(税込)'):
            ws[f'H{current_row}'] = product_info['FA売価(税込)']
            
        # Update URLs for each shop
        shop_column_mapping = {
            'waste': 'R',           # e-life＆work shop
            'kougushop': 'W',       # 工具ショップ
            'kouei-sangyou': 'AB',  # 晃栄産業　楽天市場店
            'dear-worker': 'AG'     # Dear worker ディアワーカー
        }
        
        for shop_code, column in shop_column_mapping.items():
            if shop_code in shop_info:
                shop_data = shop_info[shop_code]
                if 'URL' in shop_data:
                    ws[f'{column}{current_row}'] = shop_data['URL']
        
        current_row += 1
    
    # Save the workbook
    wb.save(excel_path)
    print(f"\nExcel file updated: {excel_path}")

if __name__ == "__main__":
    main()
    # After main() completes, update the Excel file
    update_excel_with_results('results.json', 'New folder/araki.xlsx') 
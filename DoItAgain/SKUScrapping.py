import pandas as pd
import json
import os
import requests
import time
from urllib.parse import urlparse
from collections import OrderedDict
import openpyxl
import shutil

# Rakuten API Constants
API_ENDPOINT = 'https://app.rakuten.co.jp/services/api/IchibaItem/Search/20220601'
APP_ID = '1085274442429696242'
AFFILIATE_ID = '494cbb7b.da2d8105.494cbb7c.3832fe1e'

def fetch_item_details(search_term, sku_code):
    """Fetch item details using Rakuten Ichiba Item Search API"""
    params = {
        'applicationId': APP_ID,
        'affiliateId': AFFILIATE_ID,
        'keyword': search_term,
        'hits': 30,
        'format': 'json'
    }
    
    try:
        print(f"\n=== DEBUG: Fetching data for SKU: {sku_code} ===")
        print(f"Search term: {search_term}")
        print(f"API URL: {API_ENDPOINT}")
        print(f"Params: {params}")
        
        response = requests.get(API_ENDPOINT, params=params)
        response.raise_for_status()
        
        data = response.json()
        if 'error' in data:
            print(f"API Error: {data['error']}")
            return None
            
        print(f"\nFound {len(data.get('Items', []))} items in API response")
        if data.get('Items'):
            print("\nFirst item preview:")
            first_item = data['Items'][0]['Item']
            print(f"itemName: {first_item.get('itemName', 'N/A')}")
            print(f"itemCode: {first_item.get('itemCode', 'N/A')}")
            print(f"shopName: {first_item.get('shopName', 'N/A')}")
            print(f"shopUrl: {first_item.get('shopUrl', 'N/A')}")
            print(f"itemUrl: {first_item.get('itemUrl', 'N/A')}")
            
        return data
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data: {e}")
        if hasattr(e.response, 'text'):
            print(f"Response text: {e.response.text}")
        if e.response.status_code == 429:  # Too Many Requests
            print("Rate limit hit, waiting 60 seconds...")
            time.sleep(60)
            return fetch_item_details(search_term, sku_code)  # Retry
        return None

def format_sku_for_shop(sku: str) -> str:
    """Format SKU for first23 shop"""
    try:
        # Format for first23 shop
        if sku and ':' in sku:
            return 'first23:' + sku.split(':')[1]
        return sku
    except Exception as e:
        print(f"Error formatting SKU {sku}: {e}")
        return sku

def match_shop_url(shop_name, items, search_term):
    """Match shop name with items from API response"""
    # Only look for first23 shop
    target_domain = "first23"
        
    for item in items:
        shop_url = item['Item'].get('shopUrl', '')
        if target_domain in shop_url:
            item_code = item['Item'].get('itemCode', '')
            # Format item code for first23
            formatted_code = format_sku_for_shop(item_code)
            return {
                'itemUrl': item['Item'].get('itemUrl', ''),
                'itemPrice': str(item['Item'].get('itemPrice', '')),
                'itemName': item['Item'].get('itemName', ''),
                'shopUrl': shop_url,
                'shopName': item['Item'].get('shopName', ''),
                'itemCode': formatted_code,
                '検索条件': format_sku_for_shop(search_term),  # Format search_term using the same function
                'genreId': item['Item'].get('genreId', ''),
                'tagline': item['Item'].get('tagline', ''),
                'taxIncluded': item['Item'].get('taxIncluded', False)
            }
    return None

def load_progress(json_path):
    """Load progress from existing JSON file"""
    if os.path.exists(json_path):
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return []
    return []

def excel_to_single_nested_json(excel_path, json_path=None):
    """Convert Excel to single nested JSON format"""
    if not json_path:
        json_path = os.path.splitext(excel_path)[0] + "_single.json"
    
    # Load existing progress
    all_results = load_progress(json_path)
    processed_skus = {result["shop"]["SKUコード"] for result in all_results}
    
    df = pd.read_excel(excel_path, sheet_name=0, header=2)  # Skip 2 metadata rows
    


    # Define static shop names - only first23
    shop_names = ["first23"]

    # Define column ranges for first23
    shop_columns = {
        "first23": ["価格", "ポイント", "クーポン", "在庫", "URL"]
    }
    
    try:
        # Process all rows
        for index, row in df.iterrows():
            result = {"shop": OrderedDict()}
            search_term = row.get("検索条件", "") 
            # First add SKUコード
            if "SKUコード" in df.columns:
                val = row["SKUコード"]
                result["shop"]["SKUコード"] = str(val) if pd.notna(val) else ""
            
            # Skip if already processed
            sku_code = result["shop"].get("SKUコード", "")
            if sku_code in processed_skus:
                print(f"Skipping already processed SKU {index + 1}/{len(df)}: {sku_code}")
                continue
            
            # Store API data temporarily
            api_item_name = None
            api_item_code = None
            shop_data = OrderedDict()

            # Use SKU code for initial API search
            if sku_code:
                print(f"\nSearching with SKU: {sku_code}")
                # Search in first23 shop
                api_data = fetch_item_details(sku_code, sku_code)
                if api_data and 'Items' in api_data:
                    items = api_data['Items']
                    print(f"Found {len(items)} items in API response")
                    
                    # Shop data with API information
                    for shop in shop_names:
                        shop_data[shop] = {}
                        # Get basic shop data from Excel
                        for col in shop_columns[shop]:
                            if col in df.columns:
                                val = row[col]
                                clean_col = col.split(".")[0]  # Remove suffix like ".1"
                                if pd.isna(val):
                                    shop_data[shop][clean_col] = ""
                                else:
                                    shop_data[shop][clean_col] = str(val)
                        
                        # Add API data if found
                        api_match = match_shop_url(shop, items, sku_code)
                        if api_match:
                            print(f"\nFound match in first23 shop:")
                            print(f"Item Name: {api_match['itemName']}")
                            print(f"Item Code: {api_match['itemCode']}")
                            print(f"URL: {api_match['itemUrl']}")
                            
                            # Store URL from itemUrl
                            shop_data[shop]["URL"] = api_match["itemUrl"]
                            # Store price from itemPrice
                            shop_data[shop]["価格"] = api_match["itemPrice"]
                            shop_data[shop]["shop_name"] = api_match["shopName"]
                            
                            # Store API data
                            result["shop"]["商品名"] = api_match["itemName"]
                            result["shop"]["商品管理番号"] = api_match["itemCode"]
                            result["shop"]["検索条件"] = search_term
                    
                    # Add delay to avoid API rate limits
                    time.sleep(1)

            # Now add all shop data
            for shop, data in shop_data.items():
                result["shop"][shop] = data

            all_results.append(result)
            processed_skus.add(sku_code)
            print(f"Processed SKU {index + 1}/{len(df)}: {sku_code}")
            
            # Save progress periodically
            if (index + 1) % 10 == 0:
                with open(json_path, "w", encoding="utf-8") as f:
                    json.dump(all_results, f, ensure_ascii=False, indent=2)
                print(f"Progress saved: {len(all_results)} SKUs")

    except KeyboardInterrupt:
        print("\nScript interrupted! Saving progress...")
    finally:
        # Final save
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(all_results, f, ensure_ascii=False, indent=2)
        print(f"Single nested JSON saved to: {json_path} with {len(all_results)} SKUs")

def update_excel_urls(json_path, excel_path, new_excel_path):
    """Update URLs and other data in the existing Excel file"""
    # Read JSON data
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    print("\n=== DEBUG: First item in JSON data ===")
    if data:
        print(json.dumps(data[0], indent=2, ensure_ascii=False))
    
    # Copy original Excel file
    shutil.copy2(excel_path, new_excel_path)
    
    # Load the copied Excel file
    wb = openpyxl.load_workbook(new_excel_path)
    ws = wb.active
    
    # URL cell mappings for each shop
    url_cells = {
        'e-life＆work shop': 'R',
        '工具ショップ': 'W',
        '晃栄産業　楽天市場店': 'AB',
        'Dear worker ディアワーカー': 'AG'
    }
    
    # Start from row 4
    current_row = 4
    
    # Update URLs and other data for each item
    for item in data:
        shop_data = item['shop']
        print(f"\n=== DEBUG: Processing row {current_row} ===")
        print("商品名:", shop_data.get("商品名"))
        print("商品管理番号:", shop_data.get("商品管理番号"))
        print("検索条件:", shop_data.get("検索条件"))
        
        # Get the first shop's data for common fields
        first_shop_data = None
        for shop_name in url_cells.keys():
            if shop_name in shop_data:
                shop_info = shop_data[shop_name]
                if shop_info:
                    first_shop_data = shop_info
                    print(f"\nFirst shop data from: {shop_name}")
                    print(json.dumps(first_shop_data, indent=2, ensure_ascii=False))
                    break
        
        # Update common fields if we have data
        if shop_data.get("商品名"):
            ws[f'C{current_row}'] = shop_data["商品名"]
           
        if shop_data.get("商品管理番号"):
            ws[f'B{current_row}'] = shop_data["商品管理番号"]
            
        if shop_data.get("検索条件"):
            ws[f'D{current_row}'] = shop_data["検索条件"]
            

        if first_shop_data:
            if first_shop_data.get('itemPrice'):
                ws[f'H{current_row}'] = first_shop_data['itemPrice']  # price
                print(f"Writing price to H{current_row}: {first_shop_data['itemPrice']}")
            if 'taxIncluded' in first_shop_data:
                tax_value = '1' if first_shop_data['taxIncluded'] else '0'
                ws[f'J{current_row}'] = tax_value  # tax include
                
        
        # Update shop-specific URLs
        for shop_name, column in url_cells.items():
            if shop_name in shop_data:
                shop_info = shop_data[shop_name]
                url = shop_info.get('URL', '')
                cell = f"{column}{current_row}"
                ws[cell] = url
                print(f"Writing URL for {shop_name} to {cell}: {url}")
        
        current_row += 1
    
    # Save the workbook
    wb.save(new_excel_path)
    print(f"\nURLs and data updated in Excel file: {new_excel_path}")

if __name__ == "__main__":
    excel_to_single_nested_json("araki.xlsx")
    # Update URLs in a new Excel file
    new_excel_path = "araki_with_api_" + time.strftime("%Y%m%d_%H%M%S") + ".xlsx"
    update_excel_urls("araki_single.json", "araki.xlsx", new_excel_path)

